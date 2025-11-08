from rest_framework import viewsets, mixins
from django.shortcuts import render
from rest_framework.permissions import IsAuthenticated
from rest_framework import generics
from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook

from apps.core.permissions.permissions import IsAdminOrReadOnly, IsSuperOrReadOnly
from apps.purchase.models.purchase import Purchase
from apps.purchase.serializers.purchase_serializers import PurchaseSerializer
from apps.users.models.user import User


class PurchaseViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """
    API endpoint to list, view, update, and delete loan.
    """

    permission_classes = [IsAuthenticated, IsSuperOrReadOnly, IsAdminOrReadOnly]
    queryset = Purchase.objects.filter(is_active=True)
    serializer_class = PurchaseSerializer

    def destroy(self, request, *args, **kwargs):
        purchase = self.get_object()
        purchase.is_active = False
        purchase.save(update_fields=["is_active"])
        return Response(
            {"detail": f"La compra {purchase.code} ha sido desactivado."},
            status=status.HTTP_200_OK,
        )


class PurchasesByDocumentView(APIView):
    def get(self, request, document_number):
        try:
            user = User.objects.get(document_number=document_number)
        except User.DoesNotExist:
            return Response(
                {"error": "Usuario no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        purchases = Purchase.objects.filter(user=user).prefetch_related("product_purchases__product")
        serializer = PurchaseSerializer(purchases, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class ExportPurchasesExcelView(APIView):
    def get(self, request):
        doc_type = request.GET.get("document_type")
        doc_number = request.GET.get("document_number")

        print(doc_type, doc_number)

        if not doc_type or not doc_number:
            return Response(
                {"detail": "Se requieren 'document_type' y 'document_number'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = User.objects.filter(
            document_type__short_name=doc_type, document_number=doc_number
        ).first()

        if not user:
            return Response(
                {"detail": "Usuario no encontrado."},
                status=status.HTTP_404_NOT_FOUND,
            )

        purchases = (
            Purchase.objects.filter(user=user)
            .prefetch_related("product_purchases__product")
            .order_by("-created_at")
        )

        wb = Workbook()

        ws_user = wb.active
        ws_user.title = "Cliente"
        ws_user.append(["ID", "Nombre", "Email", "Tipo Documento", "Número Documento"])
        ws_user.append([
            str(user.id),
            user.username,
            user.email,
            user.document_type.short_name if user.document_type else "",
            user.document_number or "",
        ])

        ws_purchase = wb.create_sheet(title="Compras")
        ws_purchase.append(["Código", "Total", "Fecha"])
        for p in purchases:
            ws_purchase.append([p.code, float(p.total), p.created_at.strftime("%Y-%m-%d %H:%M")])


        ws_products = wb.create_sheet(title="Productos")
        ws_products.append(["Compra", "Producto", "Precio Unitario", "Cantidad", "Subtotal"])

        for p in purchases:
            for item in p.product_purchases.all():
                ws_products.append([
                    p.code,
                    item.product.name,
                    float(item.product.price),
                    item.quantity,
                    float(item.product.price * item.quantity),
                ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="compras_{user.username}.xlsx"'

        wb.save(response)
        return response